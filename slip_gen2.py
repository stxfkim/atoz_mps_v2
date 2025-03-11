import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
from functions import *
from datetime import datetime
from datetime import date



def generate_salary_slip(df,periode):
    
    today = date.today()
    current_month = today.month
    current_year = str(today.year % 100)
    current_date = datetime.now()
    try:
        with open("last_count.txt", "r") as file:
            last_count = int(file.read())
    except FileNotFoundError:
        last_count = 0

    if current_date.day == 1:
        last_count=0
    # Copy template to new file
    

    # Get the template sheet
    

    # Get unique NIPs
    unique_nips = df["NIP"].unique()
    slips_list_filename = []
    # Process each NIP
    for nip in unique_nips:
        wb = load_workbook(Path('template/template_slip_gaji.xlsx'))
        template_sheet = wb["tmp"]
        emp_df = df[df["NIP"] == nip]  # Filter data for the NIP
        emp_name = emp_df["Nama"].iloc[0]

        # Create new sheet for this NIP
        sheet_name = f"{emp_name}_{nip}"  # Shorten name for sheet
        new_sheet = wb.copy_worksheet(template_sheet)
        new_sheet.title = sheet_name
        wb.remove(template_sheet)

        # Insert Employee Information
        new_sheet["B7"].value = emp_name
        new_sheet["B8"].value = emp_df["Jabatan"].iloc[0]
        new_sheet["B9"].value = emp_df["masa_kerja"].iloc[0]
        new_sheet["G7"].value = emp_df["Nama Bank"].iloc[0]
        new_sheet["G8"].value = emp_df["Nama Akun Bank"].iloc[0]
        new_sheet["G9"].value = emp_df["Nomor Rekening"].iloc[0]
        
        new_sheet["B28"].value = emp_df["ket_tidak_masuk"].iloc[0]
        new_sheet["B29"].value = emp_df["ket_hari_lembur"].iloc[0]
        new_sheet["B30"].value = emp_df["ket_jam_lembur"].iloc[0]
        new_sheet["B31"].value = emp_df["ket_tidak_absen"].iloc[0]
        new_sheet["I7"].value = emp_df["total_gaji_final"].iloc[0]
        new_sheet["A3"].value = "Periode: "+periode
        
        last_count=last_count+1
        bank_type = emp_df["Nama Bank"].str.upper().str.contains("BCA", na=False).map({True: "BCA", False: "NON BCA"}).iloc[0] 
        
        new_sheet["A4"].value = "No.KWT_ATS_"+int_to_roman(current_month)+"_"+current_year+"_"+str(last_count)+"_"+bank_type

        # Insert Salary Data
        start_row = 12  # Adjust based on template
        for _, row in emp_df.iterrows():
            new_sheet[f"A{start_row}"] = row["Tanggal"]
            new_sheet[f"B{start_row}"] = row["hari"]
            new_sheet[f"C{start_row}"] = row["scan_masuk"]
            new_sheet[f"D{start_row}"] = row["scan_pulang"]
            new_sheet[f"E{start_row}"] = row["jam_normal"]
            new_sheet[f"F{start_row}"] = row["jam_lembur"]
            new_sheet[f"G{start_row}"] = row["gaji_normal"]
            new_sheet[f"H{start_row}"] = row["gaji_lembur"]
            new_sheet[f"I{start_row}"] = row["uang_makan"]
            new_sheet[f"J{start_row}"] = row["total_gaji"]
            new_sheet[f"K{start_row}"] = row["total_denda"]
            new_sheet[f"L{start_row}"] = row["Kasbon"]
            new_sheet[f"M{start_row}"] = row["total_gaji_harian"]

            # Highlight holiday rows
            if row["is_holiday"] == "Y" or row["hari"] == "Minggu":
                for col in range(1, 14):
                    new_sheet[f"{get_column_letter(col)}{start_row}"].fill = PatternFill(start_color="FF9999", fill_type="solid")

            start_row += 1
        
        file_output = Path("kwitansi_output/Slip gaji_"+emp_name+".xlsx")
        wb.save(file_output)
        slips_list_filename.append(file_output)

    with open("last_count.txt", "w") as file:
        file.write(str(last_count))
    
    return slips_list_filename

