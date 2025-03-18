import pandas as pd
import numpy as np
from datetime import datetime
from datetime import date
import warnings
warnings.filterwarnings('ignore')
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
import streamlit as st
from dateutil.relativedelta import relativedelta

def check_password():
    """Returns `True` if the user had the correct password."""

    def password_entered():
        """Checks whether a password entered by the user is correct."""
        if st.session_state["password"] == st.secrets["password"]:
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # don't store password
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        # First run, show input for password.
        st.text_input(
            "Password", type="password", on_change=password_entered, key="password"
        )
        return False
    elif not st.session_state["password_correct"]:
        # Password not correct, show input + error.
        st.text_input(
            "Password", type="password", on_change=password_entered, key="password"
        )
        st.error("😕 Password incorrect")
        return False
    else:
        # Password correct.
        return True


def calculate_work_hours(row):
    if pd.isna(row["Keterangan Tidak Hadir"]) and not pd.isna(row["scan_masuk"]) and not pd.isna(row["scan_pulang"]):
        #time_delta = row["scan_pulang"] - row["scan_masuk"]
        if row["scan_masuk"] < row["Jam Masuk"]:
            time_delta = row["scan_pulang"] - row["Jam Masuk"]
        else:
            time_delta = row["scan_pulang"] - row["scan_masuk"]
        
        hours = time_delta.components.hours
        minutes = time_delta.components.minutes
        seconds = time_delta.components.seconds
        td = "{h}:{m}:{s}".format(h = hours, m = minutes,s=seconds)
        if minutes >= 50:
            hours += 1
        elif minutes >= 20:
            hours += 0.5

        if hours <= 8:
            jam_normal = hours
            jam_lembur = 0
        elif hours > 8:
            jam_normal = 8
            jam_lembur = hours - 8
        return jam_normal, jam_lembur, td
    else:
        return float('nan'), float('nan'), float('nan')

def calculate_scan_time(row):
    if row["Pulang Tengah Malam"] == "Y":
        scan_masuk= row["scan_max"]
        scan_pulang = row["scan_min"]
    else:
        scan_masuk = row["scan_min"]
        scan_pulang = row["scan_max"]
    return scan_masuk, scan_pulang


def calculate_salary(row,denda_scan_masuk,denda_scan_pulang):
    if  row["Tanggal"].weekday() == 6 or row["is_holiday"] == "Y": # tambahin kondisi kalo hari libur
        gaji_harian = (row["jam_normal"]/8) * (float(row["Gaji Harian (Pokok)"])*1.5)
        gaji_lembur = row["jam_lembur"]* (float(row["Upah Lembur"])*1.5)
    else:    
        
        gaji_harian = (row["jam_normal"]/8) * float(row["Gaji Harian (Pokok)"])
        gaji_lembur = row["jam_lembur"]*float(row["Upah Lembur"])
    total_denda_harian = ((gaji_harian+ gaji_lembur) * (row['Tidak Scan Masuk'] == 'Y') * (denda_scan_masuk / 100)) + ((gaji_harian+ gaji_lembur) * (row['Tidak Scan Pulang'] == 'Y') * (denda_scan_pulang / 100))
    total_gaji_harian = (gaji_harian + gaji_lembur + row["uang_makan"]) #- total_denda_harian
    
    return gaji_harian, gaji_lembur,total_denda_harian, total_gaji_harian,

def int_to_roman(num):
    values = [1000, 900, 500, 400, 100, 90, 50, 40, 10, 9, 5, 4, 1]
    symbols = ["M", "CM", "D", "CD", "C", "XC", "L", "XL", "X", "IX", "V", "IV", "I"]
    result = ""
    for v, s in zip(values, symbols):
        result += s * (num // v)
        num %= v
    return result

def generate_kwitansi(row):
    file_list = []
    today = date.today()

    current_month = today.month
    current_year = str(today.year % 100)
    current_date = datetime.now()
    
    try:
        with open("last_count.txt", "r") as file:
            last_count = int(file.read())
    except FileNotFoundError:
        last_count = 0

    if current_date.day == 1:
        last_count=0


    for idx,row in row.iterrows():
        wb = load_workbook("template/template_kwitansi.xlsx")
        sheet = wb.active

        #reading specific column
        # B3 - Nama
        sheet.cell(row=3, column=2).value=row["Nama"]
        sheet.merge_cells('B3:I6')
        sheet.cell(row=3, column=2).alignment = Alignment(horizontal='center', vertical='center')
        # H10 & E24 - gaji_final
        sheet.cell(row=10, column=8).value=row["gaji_final"]
        sheet.cell(row=10, column=8).alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells('H10:J10')  
        sheet.cell(row=24, column=5).value=row["gaji_final"]
        sheet.merge_cells('E24:K24')
        sheet.cell(row=24, column=5).alignment = Alignment(horizontal='left', vertical='center')
        # G28 - Nama Bank & Nama Akun Bank
        sheet.cell(row=28, column=7).value=row["Nama Bank"] + " A/n "+ row["Nama Akun Bank"]
        # G30 - Nomor Rekening
        sheet.cell(row=30, column=7).value=row["Nomor Rekening"]
        # V5 & G32 - tanggal dicetak
        sheet.cell(row=5, column=22).value= date.today().strftime('%d %b %Y')
        sheet.cell(row=32, column=7).value= date.today().strftime('%d %b %Y')
        # V3 - Nomor Kwitansi
        last_count=last_count+1
        sheet.cell(row=3, column=22).value= "KWT_ATS_"+int_to_roman(current_month)+"_"+current_year+"_"+str(last_count)
        # M14 & R14 - Periode Upah
        sheet.cell(row=14, column=13).value=row["start_date"]
        sheet.cell(row=14, column=18).value=row["end_date"]
        wb.title = row["nama_worksheet"]
        sheet.title = row["nama_worksheet"]
        file_name = "Kwitansi_"+row["nama_worksheet"]+"_"+str(row["start_date"].strftime('%d%b'))+"-"+str(row["end_date"].strftime('%d%b%Y'))+".xlsx"
        wb.save("kwitansi_output/"+file_name)
        file_list.append("kwitansi_output/"+file_name)
    with open("last_count.txt", "w") as file:
        file.write(str(last_count))
    return file_list

def check_kedisiplinan(row):
    working_hour_start = pd.Timestamp("9:00:00")
    if not pd.isna(row["Keterangan Tidak Hadir"]):
        return row["Keterangan Tidak Hadir"]
    elif row["Tidak Scan Masuk"] == "Y" or pd.isna(row["scan_masuk"]):
        return "Tidak Scan Masuk"
    elif row["Tidak Scan Pulang"] == "Y" or pd.isna(row["scan_pulang"]):
        return "Tidak Scan Pulang"



def calculate_duration(join_date):
    today = pd.to_datetime('today')
    diff = relativedelta(today, join_date)
    return f"{diff.years} Tahun {diff.months} Bulan {diff.days} Hari"


def get_periode(start_date,end_date):
    
    month_translation = {
    1: "Januari",
    2: "Februari",
    3: "Maret",
    4: "April",
    5: "Mei",
    6: "Juni",
    7: "Juli",
    8: "Agustus",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Desember"
}
    
    # start_date = datetime.strptime(start_date, "%Y-%m-%d")
    # end_date = datetime.strptime(end_date, "%Y-%m-%d")

    # Extract year, month, and day
    start_day = start_date.day
    start_month = start_date.month
    start_year = start_date.year

    end_day = end_date.day
    end_month = end_date.month
    end_year = end_date.year

    # Get month names in Indonesian
    start_month_name = month_translation[start_month]
    end_month_name = month_translation[end_month]

    # Format the output based on whether dates are in the same month or different months
    if start_year == end_year and start_month == end_month:
        # Same month
        return f"{start_day}-{end_day} {start_month_name} {start_year}"
    else:
        # Different months
        return f"{start_day} {start_month_name} - {end_day} {end_month_name} {end_year}"