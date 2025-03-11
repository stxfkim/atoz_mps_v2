import streamlit as st
import pandas as pd
import datetime
from datetime import date
from pathlib import Path
import zipfile
import warnings
from functions import *
from slip_gen2 import *
from pathlib import Path

from functions import check_password
warnings.filterwarnings("ignore")
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows




if check_password():
    st.set_page_config(page_title="Mini Payroll System", layout="wide")
    with st.sidebar:
        #    st.sidebar.title(":abacus: Mini Payroll System")

        # attendance_data = st.sidebar.file_uploader(
        #     "**Upload Data Absensi**", type=["xlsx", "xls"]
        # )

        # start_date = st.date_input("**Start Date**", date.today())
        # end_date = st.date_input("**End Date**", date.today())
        
        attendance_data = st.sidebar.file_uploader(
            "**Upload Data Absensi**", type=["xlsx", "xls"]
        )
        if attendance_data is not None:
            attendance_data_df = pd.read_excel(attendance_data)
            attendance_data_df.columns = attendance_data_df.columns.str.lower()
            attendance_data_df['tanggal'] = pd.to_datetime(attendance_data_df['tanggal'],format='%d-%m-%Y').dt.date

            min_date = attendance_data_df['tanggal'].min()
            #max_date = attendance_data_df['tanggal'].max()
            max_14_days = min_date + pd.Timedelta(days=14)
            st.write(attendance_data_df['tanggal'].min(),"-",attendance_data_df['tanggal'].max())

        else:
            min_date = date.today()
            #max_date = date.today()
            max_14_days = min_date + pd.Timedelta(days=14)

        start_date = st.date_input("**Start Date**", min_date,
                                   min_value=min_date,
                                   max_value=max_14_days)
        
        end_date = st.date_input("**End Date**", max_14_days,
                                   min_value=min_date,
                                   max_value=max_14_days)
        st.markdown("""---""")
        employee_master = st.sidebar.file_uploader(
            "**Upload Master Data Pegawai**", type=["xlsx", "xls"]
        )
        holidays_date = st.sidebar.file_uploader(
            "**Upload Data Libur & Cuti Bersama**", type=["xlsx", "xls"]
        )
        denda_scan_masuk = st.number_input("**Denda Tidak Scan Masuk (%)**", value=50, min_value=0, max_value=100)
        denda_scan_pulang = st.number_input("**Denda Tidak Scan Pulang (%)**", value=50, min_value=0, max_value=100)
        uang_makan = st.number_input("**Uang Makan Harian**", value=15000)



    tab1, tab2, tab3 = st.tabs(
        ["Raw Data ", "Hitung Gaji Harian", "Generate Report Jam Kerja"]
    )
    with tab1:

        st.write("#### Data Absensi")
        if attendance_data is not None:
            attendance_data_df = pd.read_excel(attendance_data)
            attendance_data_df["Tanggal"] = pd.to_datetime(
                attendance_data_df["Tanggal"], dayfirst=True
            )
            st.write(attendance_data_df)
        else:
            st.warning("Data Absensi belum di-upload", icon="⚠️")

        emp_master_last_updated = Path("temp_data/emp_master_last_updated.txt").read_text()
        st.write("#### Data Master Pegawai")
        st.write("last update:", emp_master_last_updated)
        if employee_master is not None:
            employee_master_df = pd.read_excel(
                employee_master, dtype={"Nomor Rekening": str}
            )
            employee_master_df.to_csv("temp_data/temp_employee_master.csv", index=None)
            f = open("temp_data/emp_master_last_updated.txt", "w")
            f.write(str(datetime.now().strftime("%Y-%m-%d %H:%M")))
            f.close()
            st.write(employee_master_df)
        else:
            my_file = Path("temp_data/temp_employee_master.csv")
            if my_file.is_file():
                st.write(pd.read_csv(my_file, dtype={"Nomor Rekening": str}))
            else:
                st.warning("Data Master Pegawai belum di-upload", icon="⚠️")

        holidays_date_last_updated = Path("temp_data/holidays_date_last_updated.txt").read_text()
        st.write("#### Data Libur & Cuti Bersama")
        st.write("last update:", holidays_date_last_updated)
        if holidays_date is not None:
            holidays_date_df = pd.read_excel(holidays_date)
            holidays_date_df.to_csv("temp_data/temp_holidays_date.csv", index=None)
            f = open("temp_data/holidays_date_last_updated.txt", "w")
            f.write(str(datetime.now().strftime("%Y-%m-%d %H:%M")))
            f.close()
            st.write(holidays_date_df)
        else:
            my_file = Path("temp_data/temp_holidays_date.csv")
            if my_file.is_file():
                st.write(pd.read_csv(my_file))
            else:
                st.warning("Data Libur & Cuti Bersama belum di-upload", icon="⚠️")


    with tab2:

        employee_master_df = pd.read_csv(
            "temp_data/temp_employee_master.csv", dtype={"Nomor Rekening": str}
        )
        holidays_date_df = pd.read_csv("temp_data/temp_holidays_date.csv")
        st.markdown(" \n\n")
        # st.write("Klik tombol dibawah untuk hitung gaji & generate kwitansi")

        col1, col2 = st.columns(2)

        with col1:
            btnHitungGaji = st.button(
                "Hitung Gaji", help="klik tombol ini untuk hitung gaji", type="primary"
            )

        if btnHitungGaji:
            with st.spinner("Loading...."):
                attendance_data_df["Tanggal"] = pd.to_datetime(
                    attendance_data_df["Tanggal"], format="%d-%m-%Y"
                )
                attendance_data_df["Tanggal"] = attendance_data_df["Tanggal"].dt.date
                filtered_attendance_df = attendance_data_df[
                    attendance_data_df["Tanggal"].between(start_date, end_date)
                ]
                # join to get payroll details
                absensi_emp_master = filtered_attendance_df.merge(
                    employee_master_df[
                        [
                            "PIN/ID",
                            "Keterangan",
                            "Gaji Harian (Pokok)",
                            "Upah Lembur",
                            "Nama Bank",
                            "Nama Akun Bank",
                            "Nomor Rekening",
                            "Jam Masuk",
                            "Jabatan",
                            "Tanggal Join"
                        ]
                    ],
                    left_on="NIP",
                    right_on="PIN/ID",
                    how="left",
                )
                absensi_emp_master["Jam Masuk"] = pd.to_datetime(
                        absensi_emp_master["Jam Masuk"], format="%H:%M:%S"
                    )
                #absensi_emp_master = absensi_emp_master[absensi_emp_master["Keterangan Tidak Hadir"].isnull()]
                
                absensi_emp_master["Tanggal"] = pd.to_datetime(
                    absensi_emp_master["Tanggal"]
                )
                holidays_date_df["Tanggal Libur"] = pd.to_datetime(
                    holidays_date_df["Tanggal Libur"]
                )
                # get holiday flag
                haris_mapping = {
                "Monday": "Senin",
                "Tuesday": "Selasa",
                "Wednesday": "Rabu",
                "Thursday": "Kamis",
                "Friday": "Jumat",
                "Saturday": "Sabtu",
                "Sunday": "Minggu"
                }
                absensi_emp_master["hari"] = absensi_emp_master["Tanggal"].dt.day_name().map(haris_mapping)

                absensi_emp_master = absensi_emp_master.merge(
                    holidays_date_df,
                    left_on="Tanggal",
                    right_on="Tanggal Libur",
                    how="left",
                )
                absensi_emp_master["is_holiday"] = (
                    absensi_emp_master["Tanggal"]
                    .isin(holidays_date_df["Tanggal Libur"])
                    .apply(lambda x: "Y" if x else "N")
                )
                absensi_emp_master = absensi_emp_master.drop(
                    columns=["PIN/ID", "Tanggal Libur"]
                )
                # st.write(absensi_emp_master)
                # get daily worker only data
                pekerja_harian = absensi_emp_master[
                    absensi_emp_master["Keterangan"] == "HARIAN"
                ]
                pekerja_harian_scan = pekerja_harian.drop(
                    columns=["Tidak Scan Masuk", "Tidak Scan Pulang"]
                )

                # convert Scan related column into datetime
                for col in list(pekerja_harian_scan.filter(regex="Scan ").columns):
                    pekerja_harian[col] = pd.to_datetime(
                        pekerja_harian_scan[col], format="%H:%M:%S"
                    )

                pekerja_harian["Tanggal"] = pd.to_datetime(pekerja_harian["Tanggal"])

                # get scan_masuk and scan_pulang
                pekerja_harian["scan_min"] = pekerja_harian[
                    list(pekerja_harian_scan.filter(regex="Scan").columns)
                ].min(axis=1)
                pekerja_harian["scan_max"] = pekerja_harian[
                    list(pekerja_harian_scan.filter(regex="Scan").columns)
                ].max(axis=1)
                
                pekerja_harian[
                    ["scan_masuk", "scan_pulang"]
                ] = pekerja_harian.apply(calculate_scan_time, axis=1, result_type="expand")
              
                # daily worker early scan (before 8AM)
                # pekerja_harian["scan_masuk"] = pekerja_harian["scan_masuk"].apply(
                #     lambda x: pd.to_datetime(pekerja_harian["Jam Masuk"], format='%I:%M:%S %p')
                #     if x <= pd.Timestamp("1900-01-01T08")
                #     else x
                # )
                pekerja_harian["scan_masuk"] = pekerja_harian["scan_masuk"].apply(
                    lambda x: pd.to_datetime(pekerja_harian["Jam Masuk"], format='%I:%M:%S %p').iloc[0]
                    if x <= pd.to_datetime(pekerja_harian["Jam Masuk"], format='%I:%M:%S %p').iloc[0]  # using an example element
                    else x
                )

                # get denda and uang makan
                # pekerja_harian["denda_tidak_scan_masuk"] = pekerja_harian[
                #     "Tidak Scan Masuk"
                # ].apply(lambda x: denda_scan_masuk if x == "Y" else 0)
                # pekerja_harian["denda_tidak_scan_pulang"] = pekerja_harian[
                #     "Tidak Scan Pulang"
                # ].apply(lambda x: denda_scan_pulang if x == "Y" else 0)
                #st.write(pekerja_harian)
                
                
                pekerja_harian['Tanggal Join'] = pd.to_datetime(pekerja_harian['Tanggal Join'])
                pekerja_harian['masa_kerja'] = pekerja_harian['Tanggal Join'].apply(calculate_duration)
                
                pekerja_harian["uang_makan"] = pekerja_harian["Uang Makan"].apply(
                    lambda x: uang_makan if x == "Y" else 0
                )
                
                # calculate working hours
                pekerja_harian[
                    ["jam_normal", "jam_lembur", "timedelta"]
                ] = pekerja_harian.apply(calculate_work_hours, axis=1, result_type="expand")
      
                pekerja_harian[["gaji_normal", "gaji_lembur","total_denda", "total_gaji"]] = pekerja_harian.apply(
                lambda row: calculate_salary(row, denda_scan_masuk, denda_scan_pulang),
                axis=1,
                result_type="expand")
                pekerja_harian["Kasbon"] = pekerja_harian["Kasbon"].fillna(0)
                pekerja_harian["total_gaji_harian"] = pekerja_harian["total_gaji"] - (pekerja_harian["total_denda"]+pekerja_harian["Kasbon"] )
                
                total_gaji_df = (
                    pekerja_harian.groupby("NIP")
                    .agg({"total_gaji": "sum","Kasbon": "sum"})
                    .rename(columns={"total_gaji": "gaji_final_sebelum_kasbon","Kasbon": "total_kasbon"})
                    .reset_index()
                )
                gaji_pekerja_harian_details = pd.merge(
                    pekerja_harian, total_gaji_df, on="NIP", how="left"
                )

                gaji_pekerja_harian_details["gaji_final"] = gaji_pekerja_harian_details["gaji_final_sebelum_kasbon"] - gaji_pekerja_harian_details["total_kasbon"] 
                
                
                df_ph_details = gaji_pekerja_harian_details.drop(columns=['Jabatan','Departemen','Kantor','scan_min','scan_max'])
                
                
                kolom_scan = df_ph_details.filter(regex='^Scan').columns.tolist()
                kolom_scan+=["scan_masuk","scan_pulang"]
                df_ph_details[kolom_scan] = df_ph_details[kolom_scan].apply(lambda x: pd.to_datetime(x).dt.strftime('%H:%M:%S'))
                df_ph_details[["Tanggal"]] = df_ph_details[["Tanggal"]].apply(lambda x: pd.to_datetime(x).dt.strftime('%d-%m-%Y'))
                
                gaji_final_df = df_ph_details.groupby("NIP").agg(
                    total_gaji_final=("total_gaji_harian", "sum")
                ).reset_index()
                
                df_ph_details = pd.merge(
                    df_ph_details, gaji_final_df, on="NIP", how="left"
                )
                
                summary_ket_slip_gaji = df_ph_details.groupby("NIP").agg(
                            ket_tidak_scan_masuk=("Tidak Scan Masuk", lambda x: (x == "Y").sum()),
                            ket_tidak_scan_pulang=("Tidak Scan Pulang", lambda x: (x == "Y").sum()),
                            ket_tidak_masuk=("Keterangan Tidak Hadir", lambda x: x.notna().sum()),  # Count non-null values
                            ket_hari_lembur=("jam_lembur", lambda x: (x > 0).sum()), 
                            ket_jam_lembur=("jam_lembur", "sum")  # Sum overtime hours
                        ).reset_index()
                
                summary_ket_slip_gaji["ket_tidak_absen"] = summary_ket_slip_gaji["ket_tidak_scan_masuk"] + summary_ket_slip_gaji["ket_tidak_scan_pulang"]
                
                summary_ket_slip_gaji_final = pd.merge(
                    summary_ket_slip_gaji,pekerja_harian[["NIP","Jabatan","masa_kerja","Nama Bank","Nama Akun Bank", "Nomor Rekening"]], on="NIP", how="left"
                )
                
                
                
                df_slip_gaji_1 = df_ph_details[["NIP","Nama","Tanggal","hari","is_holiday","scan_masuk","scan_pulang","jam_normal","jam_lembur",
                                            "gaji_normal","gaji_lembur","uang_makan","total_gaji","total_denda",
                                        "Kasbon","total_gaji_harian","total_gaji_final"]]
                
                detail_slip_gaji = pd.merge(
                    df_slip_gaji_1, summary_ket_slip_gaji_final, on="NIP", how="left"
                ).drop_duplicates()
                
                st.markdown("### Detail Gaji Pekerja Harian (preview)")
                st.dataframe(df_ph_details)
                
                
                
                st.markdown("### Gaji Pekerja Harian Summary (Slip Gaji)")
                st.dataframe(detail_slip_gaji)

                
                
                
                
                
                ph_list_nip = df_ph_details["NIP"].drop_duplicates().values.tolist()
                ph_list_filename = []
                for idx in range(len(ph_list_nip)):
                    temp_df = df_ph_details[df_ph_details["NIP"] == ph_list_nip[idx]]
                    nama_str = temp_df['Nama'].head(1).to_string(index=False).replace(' ', '_')
                    file_name = "Details_"+nama_str+"_"+str(start_date.strftime('%d%b'))+"-"+str(end_date.strftime('%d%b%Y'))+".xlsx"
                    wb = Workbook()
                    ws = wb.active
                    for r in dataframe_to_rows(temp_df, index=False, header=True):
                        ws.append(r)
                    
                    #locate the is_holiday column
                    column_header_1 = "is_holiday"
                    column_header_2 = "hari"
                    fill_color = "FF0000"

                    # Find the coordinate of the column header name
                    for row in ws.iter_rows(min_row=1, max_row=1):
                        for cell in row:
                            if cell.value == column_header_1:
                                column_index_1 = cell.column
                                break
                    for row in ws.iter_rows(min_row=1, max_row=1):
                        for cell in row:
                            if cell.value == column_header_2:
                                column_index_2 = cell.column
                                break

                    # Loop through the rows and check the value of the cell in the specified column
                    for row in ws.iter_rows(min_row=2):
                        for cell in row:
                            if (cell.column == column_index_1 and cell.value == "Y") or (cell.column == column_index_2 and cell.value == "Sunday"):
                                # Set the fill color to red
                                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                                for cell_in_row in row:
                                    cell_in_row.fill = fill
                                break
                            
                    wb.save("kwitansi_output/" + file_name)
                    
                    ph_list_filename.append("kwitansi_output/" + file_name)
                
                df_kwitansi = (
                    gaji_pekerja_harian_details[
                        [
                            "NIP",
                            "Nama",
                            "Nama Bank",
                            "Nama Akun Bank",
                            "Nomor Rekening",
                            "total_kasbon",
                            "gaji_final_sebelum_kasbon"
                        ]
                    ]
                    .drop_duplicates()
                    .reset_index(drop=True)
                )
                df_kwitansi["gaji_final"] = df_kwitansi["gaji_final_sebelum_kasbon"] - df_kwitansi["total_kasbon"] 
                df_kwitansi["start_date"] = start_date
                df_kwitansi["end_date"] = end_date
                df_kwitansi[["nama_worksheet"]] = df_kwitansi[["Nama"]].replace(
                    " ", "_", regex=True
                )
                file_list = generate_kwitansi(df_kwitansi)
                periode = get_periode(start_date,end_date)
                slips_list_filename = generate_salary_slip(detail_slip_gaji,periode)
                
                #st.write(slips_list_filename)
                
                st.markdown("### Detail Kwitansi")
                st.write(df_kwitansi)
                gaji_pekerja_harian_details.to_excel(
                    "kwitansi_output/" + "detail_perhitungan_gaji.xlsx", index=None
                )
                df_kwitansi.to_excel(
                    "kwitansi_output/" + "detail_kwitansi.xlsx", index=None
                )

                file_list.append("kwitansi_output/" + "detail_kwitansi.xlsx")
                file_list.append("kwitansi_output/" + "detail_perhitungan_gaji.xlsx")
                #file_list.append("kwitansi_output/slip_gaji_"+periode+".xlsx")
                file_list+=ph_list_filename
                file_list+=slips_list_filename
                with zipfile.ZipFile(
                    "kwitansi_output/"
                    + "Kwitansi_"
                    + str(start_date.strftime("%d%b"))
                    + "-"
                    + str(end_date.strftime("%d%b%Y"))
                    + ".zip",
                    "w",
                ) as zipMe:
                    for file in file_list:
                        zipMe.write(file, compress_type=zipfile.ZIP_DEFLATED)
        with col2:
            my_file = Path(
                "kwitansi_output/"
                + "Kwitansi_"
                + str(start_date.strftime("%d%b"))
                + "-"
                + str(end_date.strftime("%d%b%Y"))
                + ".zip"
            )
            if my_file.is_file():
                with open(
                    "kwitansi_output/"
                    + "Kwitansi_"
                    + str(start_date.strftime("%d%b"))
                    + "-"
                    + str(end_date.strftime("%d%b%Y"))
                    + ".zip",
                    "rb",
                ) as fp:
                    btn = st.download_button(
                        label="Download Kwitansi",
                        data=fp,
                        file_name="Kwitansi_"
                        + str(start_date.strftime("%d%b"))
                        + "-"
                        + str(end_date.strftime("%d%b%Y"))
                        + ".zip",
                        mime="application/zip",
                    )


    with tab3:

        col1, col2 = st.columns(2)

        with col1:
            btnGenerateReportWH = st.button(
                "Generate Report", help="klik tombol ini untuk generate report", type="primary"
            )

        if btnGenerateReportWH:
            with st.spinner("Loading...."):
                master_emp = pd.read_csv(
                    "temp_data/temp_employee_master.csv", dtype={"Nomor Rekening": str}
                )
                holidays_date_df = pd.read_csv("temp_data/temp_holidays_date.csv")
                absensi_emp_master = attendance_data_df.merge(
                                    master_emp[
                                        [
                                            "PIN/ID",
                                            "Keterangan",
                                        ]
                                    ],
                                    left_on="NIP",
                                    right_on="PIN/ID",
                                    how="left",
                                )
                absensi_emp_master["Tanggal"] = pd.to_datetime(
                                    absensi_emp_master["Tanggal"]
                                )
                holidays_date_df["Tanggal Libur"] = pd.to_datetime(
                                    holidays_date_df["Tanggal Libur"]
                                )
                haris_mapping = {
                "Monday": "Senin",
                "Tuesday": "Selasa",
                "Wednesday": "Rabu",
                "Thursday": "Kamis",
                "Friday": "Jumat",
                "Saturday": "Sabtu",
                "Sunday": "Minggu"
                }
                absensi_emp_master["hari"] = absensi_emp_master["Tanggal"].dt.day_name().map(haris_mapping)

                
            

                absensi_emp_master = absensi_emp_master.merge(
                                    holidays_date_df,
                                    left_on="Tanggal",
                                    right_on="Tanggal Libur",
                                    how="left",
                                )
                absensi_emp_master["is_holiday"] = (
                                    absensi_emp_master["Tanggal"]
                                    .isin(holidays_date_df["Tanggal Libur"])
                                    .apply(lambda x: "Y" if x else "N")
                                )
                absensi_emp_master = absensi_emp_master.drop(
                                    columns=["PIN/ID", "Tanggal Libur","Jabatan","Departemen","Kantor","PIN","Uang Makan"]
                                )
                # convert Scan related column into datetime
                all_worker_scan = absensi_emp_master.drop(columns=["Tidak Scan Masuk", "Tidak Scan Pulang"])
                for col in list(all_worker_scan.filter(regex="Scan ").columns):
                    absensi_emp_master[col] = pd.to_datetime(all_worker_scan[col], format="%H:%M:%S")

                absensi_emp_master["Tanggal"] = pd.to_datetime(absensi_emp_master["Tanggal"])

                absensi_emp_master["scan_masuk"] = absensi_emp_master[list(all_worker_scan.filter(regex="Scan").columns)].min(axis=1)

                absensi_emp_master["scan_pulang"] = absensi_emp_master[list(all_worker_scan.filter(regex="Scan").columns)].max(axis=1)
                absensi_emp_master["scan_pulang"] = absensi_emp_master.apply(
                                    lambda x: pd.NaT
                                    if x["Tidak Scan Pulang"] == "Y"
                                    else x["scan_pulang"],axis=1
                                )

                absensi_emp_master["scan_masuk"] = absensi_emp_master.apply(
                                    lambda x: pd.Timestamp("1900-01-01T09")
                                    if x["scan_masuk"] <= pd.Timestamp("1900-01-01T09") and x["Tidak Scan Masuk"] != "Y"
                                    else x["scan_masuk"],axis=1
                                )


                # daily worker early scan (before 9AM)


                absensi_emp_master[["jam_normal", "jam_lembur", "timedelta"]] = absensi_emp_master.apply(calculate_work_hours, axis=1, result_type="expand")

                absensi_emp_master["kedisiplinan"] = absensi_emp_master.apply(check_kedisiplinan, axis=1, result_type="expand")

                absensi_emp_master["total_jam_normal"] = absensi_emp_master["jam_normal"] + absensi_emp_master["jam_lembur"] 
                
                output = absensi_emp_master[["NIP","Nama","Keterangan","hari","Tanggal","scan_masuk","scan_pulang","jam_normal","jam_lembur","total_jam_normal","kedisiplinan"]]

                output.sort_values(['NIP', 'Tanggal'], inplace=True)
                st.markdown("### Detail Total Jam Kerja")
                st.dataframe(output)
                output.to_excel("report_output/" + "rekap_working_hours.xlsx",index=False)

                grouped = output.groupby(['NIP', 'Nama',"Keterangan",])["kedisiplinan"].value_counts().unstack(fill_value=0).reset_index()
                grouped.to_excel("report_output/" + "rekap_kedisiplinan.xlsx",index=False)
                st.markdown("### Detail Kedisiplinan")
                st.markdown("Kedisiplinan Periode "+ str(start_date.strftime("%d%b"))+ " -"+ str(end_date.strftime("%d%b%Y")))
                st.dataframe(grouped)
                file_list = []
                file_list.append("report_output/" + "rekap_working_hours.xlsx")
                file_list.append("report_output/" + "rekap_kedisiplinan.xlsx")
                with zipfile.ZipFile(
                    "report_output/"
                    + "Report_"
                    + str(start_date.strftime("%d%b"))
                    + "-"
                    + str(end_date.strftime("%d%b%Y"))
                    + ".zip",
                    "w",
                ) as zipMe:
                    for file in file_list:
                        zipMe.write(file, compress_type=zipfile.ZIP_DEFLATED)
        with col2:
            my_file = Path(
                "report_output/"
                + "Report_"
                + str(start_date.strftime("%d%b"))
                + "-"
                + str(end_date.strftime("%d%b%Y"))
                + ".zip"
            )
            if my_file.is_file():
                with open(
                    "report_output/"
                    + "Report_"
                    + str(start_date.strftime("%d%b"))
                    + "-"
                    + str(end_date.strftime("%d%b%Y"))
                    + ".zip",
                    "rb",
                ) as fp:
                    btn = st.download_button(
                        label="Download Report",
                        data=fp,
                        file_name="Report_"
                        + str(start_date.strftime("%d%b"))
                        + "-"
                        + str(end_date.strftime("%d%b%Y"))
                        + ".zip",
                        mime="application/zip",
                    )
        
font_css = """
    <style>
    button[data-baseweb="tab"] > div[data-testid="stMarkdownContainer"] > p {
    font-size: 20px;
    font-weight: bold;
    margin-top: 0%;
    }

    #MainMenu {visibility: hidden;}

    """

st.markdown(font_css, unsafe_allow_html=True)